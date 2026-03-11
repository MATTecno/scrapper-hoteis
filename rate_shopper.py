"""
Gera relatório Rate Shopper no formato:
  - Linhas: hotéis
  - Colunas: datas
  - Células: preço (01 PAX / 02 PAX conforme configurado)

Saída: CSV + Excel (se openpyxl instalado)
"""
import csv
import json
from datetime import datetime
from config import OUTPUT_RATE_CSV, OUTPUT_RATE_EXCEL


def montar_tabela(hoteis: list[dict]) -> dict:
    """
    Organiza os dados em:
      tabela[nome_hotel][data] = {preco, preco_original, desconto_pct, avaliacao}
    """
    tabela = {}
    datas = sorted({h["checkin"] for h in hoteis if h.get("checkin")})

    for h in hoteis:
        nome = h["nome"]
        data = h.get("checkin")
        if not data:
            continue
        if nome not in tabela:
            tabela[nome] = {"avaliacao": h.get("avaliacao"), "dados": {}}
        tabela[nome]["dados"][data] = {
            "preco":          h.get("preco_com_desconto"),
            "preco_original": h.get("preco_original"),
            "desconto_pct":   h.get("desconto_percentual"),
            "economia":       h.get("economia"),
        }

    return tabela, datas


def formatar_preco(valor):
    if valor is None:
        return "-"
    return f"R$ {valor:,.0f}".replace(",", ".")


def imprimir_rate_shopper(tabela: dict, datas: list[str]):
    """Imprime tabela no terminal."""
    col_w = 14
    nome_w = 35

    # Cabeçalho de datas
    header_datas = "  ".join(
        datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m").center(col_w)
        for d in datas
    )
    print(f"\n{'='*(nome_w + col_w * len(datas) + 10)}")
    print(f"  RATE SHOPPER — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*(nome_w + col_w * len(datas) + 10)}")
    print(f"  {'Hotel':<{nome_w}} {'Nota':>5}  {header_datas}")
    print(f"  {'-'*nome_w} {'-'*5}  {'  '.join('-'*col_w for _ in datas)}")

    for nome, info in sorted(tabela.items()):
        nota = info.get("avaliacao") or "-"
        precos = []
        for d in datas:
            dado = info["dados"].get(d)
            if dado and dado["preco"]:
                p = formatar_preco(dado["preco"])
                desc = f"(-{dado['desconto_pct']:.0f}%)" if dado.get("desconto_pct") else ""
                precos.append(f"{p} {desc}".center(col_w))
            else:
                precos.append("-".center(col_w))

        print(f"  {nome[:nome_w]:<{nome_w}} {nota:>5}  {'  '.join(precos)}")

    print(f"{'='*(nome_w + col_w * len(datas) + 10)}\n")


def salvar_csv_rate_shopper(tabela: dict, datas: list[str], caminho: str = OUTPUT_RATE_CSV):
    """Salva CSV com uma linha por hotel e colunas por data."""
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        # Cabeçalho
        cabecalho = ["Hotel", "Nota"]
        for d in datas:
            dt = datetime.strptime(d, "%Y-%m-%d")
            label = dt.strftime("%d/%m/%Y")
            cabecalho += [f"Preço {label}", f"Preço Orig {label}", f"Desc% {label}"]
        writer.writerow(cabecalho)

        # Linhas
        for nome, info in sorted(tabela.items()):
            linha = [nome, info.get("avaliacao") or ""]
            for d in datas:
                dado = info["dados"].get(d, {})
                linha += [
                    dado.get("preco") or "",
                    dado.get("preco_original") or "",
                    dado.get("desconto_pct") or "",
                ]
            writer.writerow(linha)

    print(f"CSV Rate Shopper salvo: {caminho}")


def salvar_excel_rate_shopper(tabela: dict, datas: list[str], caminho: str = OUTPUT_RATE_EXCEL):
    """Salva Excel formatado com cores. Requer openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl não instalado. Pule o Excel com: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Rate Shopper"

    # Estilos
    fill_header   = PatternFill("solid", fgColor="2D6A9F")
    fill_hotel    = PatternFill("solid", fgColor="D9E8F5")
    fill_desc     = PatternFill("solid", fgColor="C6EFCE")   # verde: tem desconto
    fill_sem_dado = PatternFill("solid", fgColor="F2F2F2")
    font_white    = Font(bold=True, color="FFFFFF")
    font_bold     = Font(bold=True)
    center        = Alignment(horizontal="center", vertical="center")
    thin          = Side(style="thin", color="BBBBBB")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Cabeçalho ──────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 7

    cabecalho_fixo = ["Hotel", "Nota"]
    for i, txt in enumerate(cabecalho_fixo, 1):
        c = ws.cell(row=1, column=i, value=txt)
        c.fill = fill_header
        c.font = font_white
        c.alignment = center
        c.border = border

    col = 3
    for d in datas:
        dt = datetime.strptime(d, "%Y-%m-%d")
        dia_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"][dt.weekday()]
        label = f"{dia_semana} {dt.strftime('%d/%m')}"

        # Mescla 3 colunas para cada data
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        c = ws.cell(row=1, column=col, value=label)
        c.fill = fill_header
        c.font = font_white
        c.alignment = center
        c.border = border

        # Sub-cabeçalho
        for j, sub in enumerate(["Preço", "Original", "Desc%"]):
            sc = ws.cell(row=2, column=col + j, value=sub)
            sc.fill = fill_header
            sc.font = font_white
            sc.alignment = center
            sc.border = border
            ws.column_dimensions[get_column_letter(col + j)].width = 12

        col += 3

    # ── Linhas de hotéis ───────────────────────────────────────────────────────
    for row_idx, (nome, info) in enumerate(sorted(tabela.items()), start=3):
        ws.cell(row=row_idx, column=1, value=nome).border = border
        ws.cell(row=row_idx, column=1).fill = fill_hotel
        ws.cell(row=row_idx, column=1).font = font_bold

        nota_cell = ws.cell(row=row_idx, column=2, value=info.get("avaliacao") or "")
        nota_cell.alignment = center
        nota_cell.border = border
        nota_cell.fill = fill_hotel

        col = 3
        for d in datas:
            dado = info["dados"].get(d, {})
            preco    = dado.get("preco")
            original = dado.get("preco_original")
            desc_pct = dado.get("desconto_pct")

            c_preco = ws.cell(row=row_idx, column=col,
                              value=preco if preco else "")
            c_preco.number_format = 'R$ #,##0'
            c_preco.alignment = center
            c_preco.border = border
            if desc_pct:
                c_preco.fill = fill_desc

            c_orig = ws.cell(row=row_idx, column=col + 1,
                             value=original if original else "")
            c_orig.number_format = 'R$ #,##0'
            c_orig.alignment = center
            c_orig.border = border

            c_desc = ws.cell(row=row_idx, column=col + 2,
                             value=f"{desc_pct:.1f}%" if desc_pct else "")
            c_desc.alignment = center
            c_desc.border = border
            if desc_pct:
                c_desc.fill = fill_desc

            if not preco:
                for j in range(3):
                    ws.cell(row=row_idx, column=col + j).fill = fill_sem_dado

            col += 3

    ws.freeze_panes = "C3"
    wb.save(caminho)
    print(f"Excel Rate Shopper salvo: {caminho}")


def gerar_rate_shopper(hoteis: list[dict]):
    tabela, datas = montar_tabela(hoteis)

    if not datas:
        print("Nenhuma data encontrada nos dados. Use scrape_rate_shopper().")
        return

    imprimir_rate_shopper(tabela, datas)
    salvar_csv_rate_shopper(tabela, datas)
    salvar_excel_rate_shopper(tabela, datas)

    # JSON completo para reuso
    with open("rate_shopper.json", "w", encoding="utf-8") as f:
        json.dump(hoteis, f, ensure_ascii=False, indent=2)
    print("JSON salvo: rate_shopper.json")
